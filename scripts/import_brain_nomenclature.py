#!/usr/bin/env python3
"""
Import brain nomenclature from Revised Brain Nomenclature.xlsx into structures/brain.yaml.

Maps Excel hierarchy (NodeID, ParentID, NodeName, ABA_ID) to BAP ontology format.
Excel NodeID 1 (Brain) maps to existing BAP_0012004. All other structures get new BAP IDs.

Usage:
    python scripts/import_brain_nomenclature.py
    python scripts/import_brain_nomenclature.py --excel path/to/file.xlsx
"""

import argparse
from pathlib import Path
from datetime import date

import yaml
from openpyxl import load_workbook

# Configuration
ROOT_DIR = Path(__file__).parent.parent
DEFAULT_EXCEL = ROOT_DIR / "Revised Brain Nomenclature.xlsx"
OUTPUT_FILE = ROOT_DIR / "structures" / "brain.yaml"
BRAIN_BAP_ID = "BAP_0012004"
BAP_ID_START = 22030  # BAP_0022030, BAP_0022031, ...


def parse_excel(filepath: Path) -> list[dict]:
    """Parse Excel file and return list of {node_id, parent_id, name, aba_id}."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header
    wb.close()

    records = []
    for row in rows:
        if not row or row[0] is None:
            continue
        node_id = int(row[0]) if row[0] is not None else None
        parent_id = int(row[1]) if row[1] is not None and str(row[1]).strip() != "" else None
        name = str(row[2]).strip() if row[2] is not None else ""
        aba_id = row[3]
        acronym = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""

        if not name:
            continue

        # Normalize parent: 0 or None -> 1 (Brain)
        if parent_id is None or parent_id == 0:
            parent_id = 1

        records.append({
            "node_id": node_id,
            "parent_id": parent_id,
            "name": name,
            "aba_id": aba_id,
            "acronym": acronym,
        })

    return records


def build_bap_mapping(records: list[dict]) -> dict[int, str]:
    """Map Excel NodeID -> BAP ID. NodeID 1 -> BAP_0012004 (Brain)."""
    mapping = {1: BRAIN_BAP_ID}
    next_bap_num = BAP_ID_START
    for r in records:
        if r["node_id"] == 1:
            continue
        mapping[r["node_id"]] = f"BAP_{next_bap_num:07d}"
        next_bap_num += 1
    return mapping


def toposort(records: list[dict]) -> list[dict]:
    """Return records in topological order (parents before children)."""
    by_node = {r["node_id"]: r for r in records}
    visited = set()
    result = []

    def visit(nid: int) -> None:
        if nid in visited:
            return
        visited.add(nid)
        parent_id = by_node[nid]["parent_id"]
        if parent_id and parent_id != nid and parent_id in by_node:
            visit(parent_id)
        if nid in by_node:
            result.append(by_node[nid])

    for r in records:
        visit(r["node_id"])

    return result


def format_aba_xref(aba_id) -> str | None:
    """Format ABA_ID as xref string, or None if invalid."""
    if aba_id is None:
        return None
    if isinstance(aba_id, str) and (aba_id.upper() == "NAN" or not aba_id.strip()):
        return None
    try:
        val = int(float(aba_id)) if isinstance(aba_id, (int, float)) else int(aba_id)
        return f"ABA:{val}"
    except (ValueError, TypeError):
        return None


def main():
    parser = argparse.ArgumentParser(description="Import brain nomenclature from Excel")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Path to Excel file")
    parser.add_argument("--output", type=Path, default=OUTPUT_FILE, help="Output YAML path")
    parser.add_argument("--dry-run", action="store_true", help="Print count only, do not write")
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"Error: Excel file not found: {args.excel}")
        return 1

    print(f"Loading {args.excel}...")
    records = parse_excel(args.excel)
    print(f"  Parsed {len(records)} rows")

    # Skip Brain (NodeID 1) - it already exists in body_regions.yaml
    brain_records = [r for r in records if r["node_id"] != 1]
    print(f"  Brain structures to add (excluding root): {len(brain_records)}")

    bap_map = build_bap_mapping(records)
    sorted_records = toposort(brain_records)

    structures = []
    for r in sorted_records:
        parent_bap = bap_map.get(r["parent_id"], BRAIN_BAP_ID)
        bap_id = bap_map[r["node_id"]]

        struct = {
            "id": bap_id,
            "name": r["name"],
            "parent": parent_bap,
            "definition": "",
        }
        xref = format_aba_xref(r["aba_id"])
        if xref:
            struct["xref"] = xref
        acronym = r.get("acronym", "").strip()
        if acronym:
            struct["abbreviation"] = acronym[:20]  # Schema maxLength 20
        structures.append(struct)

    if args.dry_run:
        print(f"Would write {len(structures)} structures to {args.output}")
        return 0

    output_data = {
        "metadata": {
            "category": "brain",
            "description": "Brain structures from Revised Brain Nomenclature (Allen Brain Atlas)",
            "version": "1.0.0",
            "last_modified": str(date.today()),
            "source": args.excel.name,
        },
        "structures": structures,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        yaml.dump(output_data, f, default_flow_style=False, allow_unicode=True, sort_keys=False)

    print(f"Wrote {len(structures)} structures to {args.output}")
    return 0


if __name__ == "__main__":
    exit(main())
